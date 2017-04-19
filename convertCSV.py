from openpyxl import load_workbook, Workbook
import csv
import os

filename = input("Enter file name without .xlsx extension: ")

workbook = load_workbook(filename + '.xlsx')

directory = filename + "_generated_files"

os.mkdir(directory)
os.chdir(directory)


def checkForBrokenRow(row):
	temp = row
	tempList = []
	for item in temp:
		tempList.append(item.offset(1,0))
	nextRow = tuple(tempList)
	if nextRow[0].value == None and nextRow[-1].value != None:
		return True
	else:
		return False

def getNextRow(row):
	temp = row
	tempList = []
	for item in temp:
		tempList.append(item.offset(1,0))
	nextRow = tuple(tempList)
	return nextRow

#go through each sheet and create a csv
for sheet in workbook:
	with open(sheet.title + '.csv', 'w', newline='') as csvfile:
		writer = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
		for row in sheet.iter_rows():
			#our normal iterator needs to skip any broken rows, we handle them separately below
			if row[0].value == None:
				continue
			#initialize cell value and row	
			currentCell = ""
			currentRow = []

			#get the next row
			nextRow = getNextRow(row)

			#adding cells to the row we're going to write
			for cell in row:
				#get rid of any quotes that currently exist in the data 
				#so there won't be any double quotes when we add them back later
				currentCell = str(cell.value).replace('"','')
				if currentCell != None and currentCell != "None":
					currentRow.append(currentCell)

			#if the next row has something in column A, go ahead and add quotes and write the current row
			if nextRow[0].value != None:
				rowToWrite = []
				for item in currentRow:
					item = '"' + str(item) + '"'
					rowToWrite.append(item)
				writer.writerow(rowToWrite)

			#if it doesn't
			if nextRow[0].value == None:
				#while the next row is still a broken row, get any data from it and concatenate to last cell of current row
				while(checkForBrokenRow(nextRow)):
					for cell in nextRow:
						nextRowCurrentCell = str(cell.value).replace('"','')
						if nextRowCurrentCell != None and nextRowCurrentCell != "None":
							currentRow[-1] = currentRow[-1] + nextRowCurrentCell
					#update nextRow so that we can keep checking if the next row is broken		
					nextRow = getNextRow(nextRow)
				#to get the last row's data, we have to go through one more time	
				for cell in nextRow:
						nextRowCurrentCell = str(cell.value).replace('"','')
						if nextRowCurrentCell != None and nextRowCurrentCell != "None":
							currentRow[-1] = currentRow[-1] + nextRowCurrentCell
				#finally, add quotes and write the row
				rowToWrite = []
				for item in currentRow:
					item = '"' + str(item) + '"'
					rowToWrite.append(item)
				writer.writerow(rowToWrite)
					

					
#create one more csv with the sheet names			
with open('sheetnames.csv', 'w', newline='') as csvfile:
	for sheet in workbook:
		writer = csv.writer(csvfile)
		currentTitle = []
		currentTitle.append(sheet.title)
		writer.writerow(currentTitle)




